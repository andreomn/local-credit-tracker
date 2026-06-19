import os
import datetime
from pathlib import Path
import json
import tempfile
import csv
import io
import re

import requests
from google.cloud import storage
import xlrd
from xlrd import xldate_as_datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# -------------------- CONFIGURAÇÕES --------------------

BASE_URL = "https://www.anbima.com.br/informacoes/merc-sec-debentures/arqs/"
MAX_DIAS_PARA_TRAS = 60

GCS_BUCKET_NAME = os.environ.get("GCS_BUCKET_NAME", "debentures-anbima-am")
GCS_PREFIX = os.environ.get("GCS_PREFIX", "anbima_debentures/")

CRICRA_URL = "https://data.anbima.com.br/busca/certificado-de-recebiveis?view=precos&page=0&q=&size="
CRICRA_GCS_PREFIX = os.environ.get("CRICRA_GCS_PREFIX", "anbima_cricra/")

MESES_NUM_PARA_TXT = {
    1: "jan", 2: "fev", 3: "mar", 4: "abr",
    5: "mai", 6: "jun", 7: "jul", 8: "ago",
    9: "set", 10: "out", 11: "nov", 12: "dez",
}

MESES_TXT_PARA_NUM = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4,
    "mai": 5, "jun": 6, "jul": 7, "ago": 8,
    "set": 9, "out": 10, "nov": 11, "dez": 12,
}

SHEETS_TO_READ = [
    ("DI_SPREAD", "CDI"),
    ("IPCA_SPREAD", "IPCA"),
]


# -------------------- HELPERS --------------------

def nome_arquivo_para_data_download(d: datetime.date) -> str:
    dia = d.day
    mes_txt = MESES_NUM_PARA_TXT[d.month]
    ano_2 = str(d.year)[-2:]
    return f"d{ano_2}{mes_txt}{dia:02d}.xls"


def nome_xlsx_cricra_para_data(d: datetime.date) -> str:
    dia = d.day
    mes_txt = MESES_NUM_PARA_TXT[d.month]
    ano_2 = str(d.year)[-2:]
    return f"d{ano_2}{mes_txt}{dia:02d}.xlsx"


def normalizar_header_coluna(valor) -> str:
    if valor is None:
        return ""

    texto = str(valor).strip().lower()

    substituicoes = {
        "á": "a", "à": "a", "â": "a", "ã": "a",
        "é": "e", "ê": "e", "è": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
        "ª": "a",
        "º": "o",
    }

    for origem, destino in substituicoes.items():
        texto = texto.replace(origem, destino)

    texto = texto.replace("%", " pct ")
    texto = texto.replace("/", " ")
    texto = texto.replace("-", " ")
    texto = texto.replace("_", " ")
    texto = " ".join(texto.split())

    return texto


def criar_mapa_headers(headers):
    mapa = {}

    for idx, header in enumerate(headers):
        h = normalizar_header_coluna(header)

        if h and h not in mapa:
            mapa[h] = idx

    return mapa


def encontrar_coluna_por_candidatos(headers, candidatos):
    mapa = criar_mapa_headers(headers)

    for candidato in candidatos:
        c = normalizar_header_coluna(candidato)

        if c in mapa:
            return mapa[c]

    return None


def valor_linha(row, idx):
    if idx is None:
        return None

    if idx < 0 or idx >= len(row):
        return None

    return row[idx]


def parse_data_flexivel(valor: str):
    if not valor:
        return None

    v = str(valor).strip()

    if not v:
        return None

    v10 = v[:10]

    formatos = [
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d.%m.%Y",
    ]

    for fmt in formatos:
        try:
            return datetime.datetime.strptime(v10, fmt).date()
        except Exception:
            pass

    return None


def parse_numero_brasileiro(valor):
    if valor is None:
        return None

    v = str(valor).strip()

    if not v:
        return None

    v = v.replace("%", "").strip()

    chars_validos = set("0123456789.,-")

    if any(c not in chars_validos for c in v):
        return str(valor).strip()

    if not any(c.isdigit() for c in v):
        return str(valor).strip()

    try:
        v_convertido = v.replace(".", "").replace(",", ".")
        return float(v_convertido)
    except Exception:
        return str(valor).strip()


def excel_serial_date_to_date(valor):
    numero = parse_numero_brasileiro(valor)

    if not isinstance(numero, (int, float)):
        return None

    if numero < 25000 or numero > 60000:
        return None

    try:
        base = datetime.date(1899, 12, 30)
        return base + datetime.timedelta(days=int(numero))
    except Exception:
        return None


def parse_data_openpyxl(valor):
    if valor is None:
        return None

    if isinstance(valor, datetime.datetime):
        return valor.date().isoformat()

    if isinstance(valor, datetime.date):
        return valor.isoformat()

    dt = parse_data_flexivel(str(valor).strip())

    if dt:
        return dt.isoformat()

    dt = excel_serial_date_to_date(valor)

    if dt:
        return dt.isoformat()

    return None


def parse_float_cell(value):
    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        v = value.replace("%", "").replace(",", ".").strip()

        if not v:
            return None

        try:
            return float(v)
        except ValueError:
            return None

    return None


def parse_float_generico(valor):
    if valor is None:
        return None

    if isinstance(valor, (int, float)):
        try:
            if isinstance(valor, float) and valor != valor:
                return None

            return float(valor)
        except Exception:
            return None

    return parse_float_cell(valor)


def parse_texto_generico(valor):
    if valor is None:
        return None

    texto = str(valor).strip()

    return texto or None


def normalizar_index_cricra(valor):
    texto = parse_texto_generico(valor)

    if not texto:
        return "prefixado"

    texto_limpo = " ".join(str(texto).strip().split())
    texto_norm = normalizar_header_coluna(texto_limpo)

    # CRI/CRA pode vir como "DI + 0,75%" no XLSX da ANBIMA.
    # Para alinhar com debêntures e com os filtros do app, tratamos DI como CDI.
    if texto_norm.startswith("cdi") or texto_norm.startswith("di"):
        return "CDI"

    if texto_norm.startswith("ipca"):
        return "IPCA"

    if texto_norm.startswith("igpm") or texto_norm.startswith("igp m"):
        return "IGP-M"

    if texto_norm in ("pre", "prefixado", "prefixada", "taxa prefixada"):
        return "prefixado"

    return texto_limpo


def parse_data_de_nome_arquivo_flexivel(nome_arquivo: str):
    nome = Path(nome_arquivo).stem.lower()
    match = re.search(r"d(\d{2})([a-z]{3})(\d{2})", nome)

    if not match:
        raise ValueError(f"Nome de arquivo inesperado: {nome_arquivo}")

    ano_2 = int(match.group(1))
    mes_abrev = match.group(2)
    dia = int(match.group(3))

    ano = 2000 + ano_2
    mes = MESES_TXT_PARA_NUM[mes_abrev]

    return datetime.date(ano, mes, dia)


# -------------------- DOWNLOAD DEBÊNTURES --------------------

def encontrar_ultimo_arquivo():
    hoje = datetime.date.today()

    for i in range(MAX_DIAS_PARA_TRAS):
        d = hoje - datetime.timedelta(days=i)
        nome = nome_arquivo_para_data_download(d)
        url = BASE_URL + nome

        print(f"Tentando baixar: {url}")

        try:
            resp = requests.get(url, timeout=30)
        except Exception as e:
            print(f"Erro de rede em {url}: {e}")
            continue

        if resp.status_code == 200:
            print(f"Encontrado arquivo para download: {nome}")
            return d, nome, resp.content

    print("AVISO: Nenhuma planilha nova encontrada nos últimos dias.")
    return None, None, None


def upload_xls_para_bucket(nome_arquivo: str, conteudo: bytes):
    client = storage.Client()
    bucket = client.bucket(GCS_BUCKET_NAME)

    prefix = GCS_PREFIX.rstrip("/") + "/"
    blob_name = prefix + nome_arquivo

    print(f"Fazendo upload para gs://{GCS_BUCKET_NAME}/{blob_name}")

    blob = bucket.blob(blob_name)
    blob.upload_from_string(conteudo, content_type="application/vnd.ms-excel")

    print("Upload concluído.")


# -------------------- CRI/CRA DOWNLOAD + CONVERSÃO --------------------

def decode_csv_anbima(conteudo: bytes) -> str:
    encodings = ["utf-8-sig", "cp1252", "latin1", "utf-8"]

    melhor_texto = None
    melhor_score = -1
    melhor_encoding = None

    for enc in encodings:
        try:
            texto = conteudo.decode(enc, errors="replace")
        except Exception:
            continue

        score = 0
        score -= texto.count(" ") * 100
        score -= texto.count("□") * 100
        score -= texto.count("Ã") * 20
        score -= texto.count("Â") * 20

        for ch in "áàâãéêíóôõúçÁÀÂÃÉÊÍÓÔÕÚÇªº":
            score += texto.count(ch)

        lower = texto.lower()

        for termo in ["referência", "referencia", "crédito", "credito", "emissão", "emissao", "código", "codigo"]:
            if termo in lower:
                score += 50

        if score > melhor_score:
            melhor_score = score
            melhor_texto = texto
            melhor_encoding = enc

    print(f"Encoding escolhido para CSV CRI/CRA: {melhor_encoding}")

    return melhor_texto or conteudo.decode("latin1", errors="replace")


def detectar_delimitador_csv(texto: str) -> str:
    linhas = texto.splitlines()
    primeira_linha = linhas[0] if linhas else ""

    if primeira_linha.count(";") >= primeira_linha.count(","):
        return ";"

    return ","


def header_eh_data(header: str) -> bool:
    h = (header or "").strip().lower()

    return (
        "data" in h
        or "vencimento" in h
        or "referência" in h
        or "referencia" in h
    )


def tentar_parsear_data_coluna(valor, header: str):
    if not header_eh_data(header):
        return None

    dt = parse_data_flexivel(str(valor).strip())

    if dt:
        return dt

    dt = excel_serial_date_to_date(valor)

    if dt:
        return dt

    return None


def tentar_extrair_ultima_data_do_csv(conteudo: bytes):
    texto = decode_csv_anbima(conteudo)
    delimitador = detectar_delimitador_csv(texto)

    reader = csv.DictReader(io.StringIO(texto), delimiter=delimitador)

    if not reader.fieldnames:
        return None

    colunas_referencia = []

    for c in reader.fieldnames:
        if not c:
            continue

        c_norm = normalizar_header_coluna(c)

        if "data" in c_norm and "referencia" in c_norm:
            colunas_referencia.append(c)

    datas = []

    for row in reader:
        for col in colunas_referencia:
            valor = (row.get(col) or "").strip()
            dt = tentar_parsear_data_coluna(valor, col)

            if dt:
                datas.append(dt)

    if not datas:
        return None

    return max(datas)


def csv_cricra_para_xlsx(conteudo_csv: bytes) -> bytes:
    texto = decode_csv_anbima(conteudo_csv)
    delimitador = detectar_delimitador_csv(texto)

    reader = csv.reader(io.StringIO(texto), delimiter=delimitador)

    wb = Workbook()
    ws = wb.active
    ws.title = "CRI_CRA"

    headers = []

    for row_idx, row in enumerate(reader, start=1):
        if row_idx == 1:
            headers = [str(v).strip() if v is not None else "" for v in row]

        for col_idx, valor in enumerate(row, start=1):
            valor_limpo = str(valor).strip() if valor is not None else ""

            if row_idx == 1:
                ws.cell(row=row_idx, column=col_idx, value=valor_limpo)
            else:
                header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ""

                dt = tentar_parsear_data_coluna(valor_limpo, header)

                if dt:
                    cell = ws.cell(row=row_idx, column=col_idx, value=dt)
                    cell.number_format = "DD/MM/YYYY"
                    continue

                valor_convertido = parse_numero_brasileiro(valor_limpo)

                ws.cell(row=row_idx, column=col_idx, value=valor_convertido)

    ws.freeze_panes = "A2"

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for cell in ws[col_letter]:
            if cell.value is not None:
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    tamanho = 10
                else:
                    tamanho = len(str(cell.value))

                max_len = max(max_len, tamanho)

        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    output = io.BytesIO()
    wb.save(output)

    return output.getvalue()


def upload_cricra_excel_para_bucket(nome_arquivo: str, conteudo: bytes):
    client = storage.Client()
    bucket = client.bucket(GCS_BUCKET_NAME)

    prefix = CRICRA_GCS_PREFIX.rstrip("/") + "/"
    blob_name = prefix + nome_arquivo

    print(f"Fazendo upload do Excel CRI/CRA para gs://{GCS_BUCKET_NAME}/{blob_name}")

    blob = bucket.blob(blob_name)

    blob.upload_from_string(
        conteudo,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    print("Upload do Excel CRI/CRA concluído.")


def baixar_cricra_csv_converter_excel_para_bucket():
    print("Iniciando download do CSV de CRI/CRA da ANBIMA Data...")

    try:
        from playwright.sync_api import sync_playwright
    except Exception as e:
        raise RuntimeError(
            "Playwright não está instalado. Adicione 'playwright' no requirements.txt "
            "e instale o Chromium no Dockerfile com: "
            "RUN python -m playwright install --with-deps chromium"
        ) from e

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )

        page = browser.new_page(accept_downloads=True)

        try:
            print(f"Abrindo página CRI/CRA: {CRICRA_URL}")

            page.goto(CRICRA_URL, wait_until="networkidle", timeout=120_000)

            for texto_botao in ["Aceitar", "Aceito", "Entendi", "Fechar", "OK"]:
                try:
                    page.get_by_text(texto_botao, exact=True).first.click(timeout=2_000)
                    print(f"Banner/modal fechado com botão: {texto_botao}")
                    break
                except Exception:
                    pass

            print("Procurando botão/link CSV...")

            candidatos = [
                page.get_by_text("CSV", exact=True).first,
                page.locator("button:has-text('CSV')").first,
                page.locator("a:has-text('CSV')").first,
                page.locator("a[href*='csv']").first,
                page.locator("[download]").first,
            ]

            download = None
            ultimo_erro = None

            for candidato in candidatos:
                try:
                    with page.expect_download(timeout=120_000) as download_info:
                        candidato.click(timeout=20_000)

                    download = download_info.value
                    break
                except Exception as e:
                    ultimo_erro = e
                    continue

            if download is None:
                raise RuntimeError(f"Não consegui acionar o download CSV. Último erro: {ultimo_erro}")

            suggested_name = download.suggested_filename or ""

            print(f"Download iniciado. Nome sugerido pela ANBIMA: {suggested_name}")

            with tempfile.TemporaryDirectory() as tmpdir:
                caminho = os.path.join(tmpdir, suggested_name or "cricra.csv")
                download.save_as(caminho)

                with open(caminho, "rb") as f:
                    conteudo_csv = f.read()

            if not conteudo_csv or len(conteudo_csv) < 100:
                raise RuntimeError("CSV CRI/CRA veio vazio ou pequeno demais.")

            amostra = decode_csv_anbima(conteudo_csv[:1000]).lower()

            if "<html" in amostra or "<!doctype html" in amostra:
                raise RuntimeError("O arquivo baixado parece ser HTML, não CSV.")

            data_csv = tentar_extrair_ultima_data_do_csv(conteudo_csv)

            if data_csv:
                nome_final = nome_xlsx_cricra_para_data(data_csv)
                print(f"Última data encontrada no CSV CRI/CRA: {data_csv.isoformat()}")
            else:
                hoje = datetime.date.today()
                nome_final = nome_xlsx_cricra_para_data(hoje)

                print(
                    "Não consegui identificar a data dentro do CSV CRI/CRA. "
                    f"Usando a data de hoje no nome: {hoje.isoformat()}"
                )

            print("Convertendo CSV CRI/CRA para Excel normalizado...")

            conteudo_xlsx = csv_cricra_para_xlsx(conteudo_csv)

            upload_cricra_excel_para_bucket(nome_final, conteudo_xlsx)

            print(f"Excel CRI/CRA salvo como {nome_final}")

        finally:
            browser.close()


# -------------------- HISTÓRICO DEBÊNTURES --------------------

def parse_data_de_nome_arquivo(nome_arquivo: str) -> datetime.date:
    stem = Path(nome_arquivo).stem.lower()
    s = stem[1:]

    if len(s) < 7:
        raise ValueError(f"Nome de arquivo inesperado: {nome_arquivo}")

    ano_2 = int(s[:2])
    mes_abrev = s[2:5]
    dia = int(s[5:7])

    ano = 2000 + ano_2
    mes = MESES_TXT_PARA_NUM[mes_abrev]

    return datetime.date(ano, mes, dia)


def parse_maturity_cell(value, datemode):
    if isinstance(value, (int, float)):
        try:
            dt = xldate_as_datetime(value, datemode)
            return dt.date().isoformat()
        except Exception:
            return str(value).strip()

    if isinstance(value, str):
        v = value.strip()
        return v or None

    return None


# -------------------- CONSOLIDAÇÃO CRI/CRA --------------------

def gerar_historico_cricra_no_mesmo_json(bucket, historico):
    cricra_prefix = CRICRA_GCS_PREFIX.rstrip("/") + "/"

    blobs = list(bucket.list_blobs(prefix=cricra_prefix))

    print(f"Encontrados {len(blobs)} blobs sob prefixo {cricra_prefix} para CRI/CRA")

    for blob in blobs:
        nome = blob.name.split("/")[-1]

        if not nome.lower().endswith(".xlsx"):
            continue

        try:
            data_arquivo = parse_data_de_nome_arquivo_flexivel(nome)
        except Exception as e:
            print(f"Ignorando CRI/CRA {blob.name} (erro ao parsear data do nome: {e})")
            continue

        print(f"Processando CRI/CRA {blob.name} (data arquivo {data_arquivo.isoformat()})")

        try:
            conteudo = blob.download_as_bytes()

            wb = load_workbook(
                io.BytesIO(conteudo),
                data_only=True,
                read_only=True,
            )

        except Exception as e:
            print(f"Erro ao abrir CRI/CRA {blob.name}: {e}")
            continue

        try:
            if "CRI_CRA" in wb.sheetnames:
                ws = wb["CRI_CRA"]
            else:
                ws = wb[wb.sheetnames[0]]

            rows_iter = ws.iter_rows(values_only=True)

            try:
                headers_raw = next(rows_iter)
            except StopIteration:
                print(f"Ignorando CRI/CRA {blob.name}: planilha vazia")
                continue

            headers = [str(v).strip() if v is not None else "" for v in headers_raw]

            col_data_ref = encontrar_coluna_por_candidatos(
                headers,
                ["Data de Referência", "Data de Referencia"],
            )
            col_risco = encontrar_coluna_por_candidatos(
                headers,
                ["Risco de Crédito", "Risco de Credito"],
            )
            col_serie = encontrar_coluna_por_candidatos(headers, ["Série", "Serie"])
            col_emissao = encontrar_coluna_por_candidatos(headers, ["Emissão", "Emissao"])
            col_codigo = encontrar_coluna_por_candidatos(headers, ["Código", "Codigo"])
            col_venc = encontrar_coluna_por_candidatos(headers, ["Vencimento"])
            col_index = encontrar_coluna_por_candidatos(
                headers,
                ["Índice / Correção", "Indice / Correcao", "Índice", "Indice"],
            )
            col_taxa_compra = encontrar_coluna_por_candidatos(headers, ["Taxa Compra"])
            col_taxa_venda = encontrar_coluna_por_candidatos(headers, ["Taxa Venda"])
            col_taxa_indicativa = encontrar_coluna_por_candidatos(headers, ["Taxa Indicativa"])
            col_desvio = encontrar_coluna_por_candidatos(headers, ["Desvio Padrão", "Desvio Padrao"])
            col_pu = encontrar_coluna_por_candidatos(headers, ["PU"])
            col_pu_par = encontrar_coluna_por_candidatos(
                headers,
                ["% PU Par / % VNE", "PU Par / VNE", "% PU Par", "% VNE"],
            )
            col_duration = encontrar_coluna_por_candidatos(headers, ["Duration"])
            col_ntnb = encontrar_coluna_por_candidatos(headers, ["Referência NTNB", "Referencia NTNB"])
            col_reune = encontrar_coluna_por_candidatos(headers, ["% Reune", "Reune"])

            if col_codigo is None:
                print(f"Ignorando CRI/CRA {blob.name}: coluna Código não encontrada.")
                print(f"Headers encontrados: {headers}")
                continue

            registros_lidos = 0

            for row in rows_iter:
                codigo = parse_texto_generico(valor_linha(row, col_codigo))

                if not codigo:
                    continue

                data_ref = parse_data_openpyxl(valor_linha(row, col_data_ref))

                if not data_ref:
                    data_ref = data_arquivo.isoformat()

                taxa_indicativa = parse_float_generico(valor_linha(row, col_taxa_indicativa))
                pu = parse_float_generico(valor_linha(row, col_pu))
                pu_par = parse_float_generico(valor_linha(row, col_pu_par))

                taxa_compra = parse_float_generico(valor_linha(row, col_taxa_compra))
                taxa_venda = parse_float_generico(valor_linha(row, col_taxa_venda))
                desvio_padrao = parse_float_generico(valor_linha(row, col_desvio))
                duration = parse_float_generico(valor_linha(row, col_duration))
                reune_pct = parse_float_generico(valor_linha(row, col_reune))

                if (
                    taxa_indicativa is None
                    and pu is None
                    and pu_par is None
                    and taxa_compra is None
                    and taxa_venda is None
                ):
                    continue

                indexador = normalizar_index_cricra(valor_linha(row, col_index))

                # IMPORTANTE:
                # Para CRI/CRA, o issuer vem da coluna "Risco de Crédito",
                # não da coluna "Emissor".
                emissor = parse_texto_generico(valor_linha(row, col_risco))

                vencimento = parse_data_openpyxl(valor_linha(row, col_venc))
                serie = parse_texto_generico(valor_linha(row, col_serie))
                emissao = parse_texto_generico(valor_linha(row, col_emissao))
                referencia_ntnb = parse_texto_generico(valor_linha(row, col_ntnb))

                historico.setdefault(codigo, []).append(
                    {
                        "date": data_ref,
                        "index": indexador,
                        "issuer": emissor,
                        "maturity_date": vencimento,
                        "taxa_indicativa": taxa_indicativa,
                        "pu": pu,
                        "brl_cents": pu_par,
                        "type": "cri/cra",
                        "series": serie,
                        "issuance": emissao,
                        "taxa_compra": taxa_compra,
                        "taxa_venda": taxa_venda,
                        "desvio_padrao": desvio_padrao,
                        "duration": duration,
                        "referencia_ntnb": referencia_ntnb,
                        "reune_pct": reune_pct,
                    }
                )

                registros_lidos += 1

            print(f"  CRI/CRA adicionados de {blob.name}: {registros_lidos} registros")

        finally:
            try:
                wb.close()
            except Exception:
                pass


# -------------------- CONSOLIDAÇÃO PRINCIPAL --------------------

def gerar_historico_pupar():
    client = storage.Client()
    bucket = client.bucket(GCS_BUCKET_NAME)

    prefix = GCS_PREFIX.rstrip("/") + "/"

    blobs = list(bucket.list_blobs(prefix=prefix))

    print(f"Encontrados {len(blobs)} blobs sob prefixo {prefix}")

    historico = {}

    for blob in blobs:
        nome = blob.name.split("/")[-1]

        if not nome.lower().endswith(".xls"):
            continue

        try:
            data_arquivo = parse_data_de_nome_arquivo(nome)
        except Exception as e:
            print(f"Ignorando {blob.name} (erro ao parsear data: {e})")
            continue

        print(f"Processando {blob.name} (data {data_arquivo.isoformat()})")

        try:
            conteudo = blob.download_as_bytes()
            book = xlrd.open_workbook(file_contents=conteudo)
        except Exception as e:
            print(f"Erro ao abrir {blob.name}: {e}")
            continue

        datemode = book.datemode

        for sheet_name, index_label in SHEETS_TO_READ:
            try:
                sheet = book.sheet_by_name(sheet_name)
            except xlrd.biffh.XLRDError:
                continue

            print(f"  Lendo aba {sheet_name} ({index_label})")

            COL_CODIGO = 0
            COL_EMISSOR = 1
            COL_VENC = 2
            COL_TAXA = 6
            COL_PU = 10
            COL_PU_PAR = 11

            for row_idx in range(1, sheet.nrows):
                cod_val = sheet.cell_value(row_idx, COL_CODIGO)
                codigo = str(cod_val).strip()

                if not codigo:
                    continue

                emissor_val = sheet.cell_value(row_idx, COL_EMISSOR)
                venc_val = sheet.cell_value(row_idx, COL_VENC)
                taxa_val = sheet.cell_value(row_idx, COL_TAXA)
                pu_val = sheet.cell_value(row_idx, COL_PU)
                pu_par_val = sheet.cell_value(row_idx, COL_PU_PAR)

                emissor = str(emissor_val).strip() if emissor_val not in ("", None) else None
                vencimento = parse_maturity_cell(venc_val, datemode)
                taxa = parse_float_cell(taxa_val)
                pu = parse_float_cell(pu_val)
                pu_par = parse_float_cell(pu_par_val)

                if taxa is None and pu is None and pu_par is None:
                    continue

                historico.setdefault(codigo, []).append(
                    {
                        "date": data_arquivo.isoformat(),
                        "index": index_label,
                        "issuer": emissor,
                        "maturity_date": vencimento,
                        "taxa_indicativa": taxa,
                        "pu": pu,
                        "brl_cents": pu_par,
                        "type": "debenture",
                    }
                )

    try:
        gerar_historico_cricra_no_mesmo_json(bucket, historico)
    except Exception as e:
        print("AVISO: Erro ao consolidar CRI/CRA no JSON.")
        print("Esse erro NÃO vai quebrar a consolidação atual de debêntures.")
        print(e)

    for codigo, lista in historico.items():
        lista.sort(key=lambda x: x["date"] or "")

    json_str = json.dumps(historico, ensure_ascii=False)

    dest_blob_name = prefix + "historico-anbima.json"
    dest_blob = bucket.blob(dest_blob_name)

    dest_blob.upload_from_string(json_str, content_type="application/json")

    print(f"Histórico consolidado salvo em gs://{GCS_BUCKET_NAME}/{dest_blob_name}")
    print(f"Total de códigos diferentes: {len(historico)}")


# -------------------- MAIN --------------------

def main():
    try:
        data, nome, conteudo = encontrar_ultimo_arquivo()

        if nome is None or conteudo is None:
            print("Nenhum arquivo novo para subir hoje. Usando apenas histórico existente.")
        else:
            upload_xls_para_bucket(nome, conteudo)

    except Exception as e:
        print("ERRO REAL ao tentar baixar ou subir planilha:")
        print(e)
        raise

    try:
        baixar_cricra_csv_converter_excel_para_bucket()
    except Exception as e:
        print("AVISO: Erro ao baixar/converter CRI/CRA.")
        print("Esse erro NÃO vai quebrar o fluxo atual de debêntures.")
        print(e)

    try:
        gerar_historico_pupar()
    except Exception as e:
        print("ERRO na geração do histórico consolidado:")
        print(e)
        raise

    print("Job finalizado com sucesso.")


if __name__ == "__main__":
    main()